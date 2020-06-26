import openpyxl


wb = openpyxl.Workbook()
lst_sheets = wb.sheetnames
print(lst_sheets[0])
sheetName = lst_sheets[0]
mySheet = wb[sheetName]

mySheet['A1'] = 'writing in Excel'
mySheet.cell(2, 2).value = 'Learning Excel'

wb.save('newlyCreatedExcel.xlsx')

wb.create_sheet('newSheet')

wb.save('newlyCreatedExcel.xlsx')

wb.create_sheet(index=0, title='moreSheet')
wb.save('newlyCreatedExcel.xlsx')
wb.close()
