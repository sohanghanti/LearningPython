import openpyxl


# open existing excel file
# automatically gets the current working directory for the path of the file
workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))
print(workbook)

# get the names of the sheets
lst_sheets = workbook.sheetnames
print(len(lst_sheets))
print(lst_sheets)

# read a particular sheet
sht1 = lst_sheets[0]
print(sht1)
mySheet = workbook[sht1]

print(type(mySheet))
row_count = mySheet.max_row
column_count = mySheet.max_column
print(row_count)
print(column_count)
# read the value from particular cell
print(mySheet['A1'].value)

# read the cell value by passing row and column
print(mySheet.cell(1, 2).value)


for i in range(1, 4):
    print(mySheet.cell(i, 2).value)
